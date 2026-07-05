# -*- coding: utf-8 -*-
"""
Dashboard Excel — Observatorio de Deforestación CORPOURABA (2000-2024)
======================================================================
Genera un libro XLSX autónomo y dinámico (fórmulas nativas, sin macros):

  Portada · Dashboard (KPIs + gráficos) · Consulta municipio (desplegable)
  Matriz municipio×periodo · Serie regional · Serie municipal (datos)
  Aux (cálculos) · Diccionario y metodología
  [+ Dinámica bosque y Hallazgos si la investigación ya generó sus salidas]

Todos los agregados son FÓRMULAS de Excel sobre la hoja de datos, de modo que
un ingeniero pueda auditar y extender el libro sin herramientas adicionales.

Uso:  python etl/build_excel_dashboard.py
Salida: entregables/Dashboard_Deforestacion_CORPOURABA_2000-2024.xlsx
"""
import json
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parent.parent
PROC = ROOT / "data" / "processed"
OUT_DIR = ROOT / "entregables"
OUT = OUT_DIR / "Dashboard_Deforestacion_CORPOURABA_2000-2024.xlsx"

# ── estilos ──────────────────────────────────────────────────────────────────
VERDE_OSC = "0B3D25"
VERDE = "1F7347"
AMBAR = "F97316"
CREMA = "FEF3C7"
GRIS = "F3F4F6"
F_TITULO = Font(name="Calibri", size=22, bold=True, color=VERDE_OSC)
F_SUB = Font(name="Calibri", size=12, color="374151")
F_H1 = Font(name="Calibri", size=14, bold=True, color=VERDE_OSC)
F_HEAD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
F_KPI_NUM = Font(name="Calibri", size=20, bold=True, color=VERDE)
F_KPI_LBL = Font(name="Calibri", size=10, color="6B7280")
F_NOTA = Font(name="Calibri", size=9, italic=True, color="6B7280")
FILL_HEAD = PatternFill("solid", start_color=VERDE)
FILL_KPI = PatternFill("solid", start_color=GRIS)
FILL_EST = PatternFill("solid", start_color=CREMA)
BORDE = Border(*[Side(style="thin", color="D1D5DB")] * 4)
CENTRO = Alignment(horizontal="center", vertical="center")
NUM_HA = "#,##0"
NUM_HA1 = "#,##0.0"
NUM_PCT = "0.0%"


def head(ws, fila, textos, anchos=None):
    for i, t in enumerate(textos, start=1):
        c = ws.cell(row=fila, column=i, value=t)
        c.font, c.fill, c.alignment, c.border = F_HEAD, FILL_HEAD, CENTRO, BORDE
    if anchos:
        for i, a in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = a


def main() -> int:
    OUT_DIR.mkdir(exist_ok=True)
    serie = pd.read_csv(PROC / "serie_municipal.csv", encoding="utf-8-sig",
                        dtype={"codigo_dane": str})
    serie["estimado"] = serie["estimado"].astype(str).str.lower().isin(["true", "1"])
    reg = pd.read_csv(PROC / "serie_regional.csv", encoding="utf-8-sig")
    reg["estimado"] = reg["estimado"].astype(str).str.lower().isin(["true", "1"])
    reg_d = (reg[reg["clase"] == "Deforestación"]
             .sort_values("ano_inicio").reset_index(drop=True))
    meta = json.loads((PROC / "metadata.json").read_text(encoding="utf-8"))
    municipios = sorted(serie["municipio"].unique())
    periodos = [p["id"] for p in meta["periodos"]]
    n_datos = len(serie)
    FIN = n_datos + 1  # última fila de la hoja de datos

    wb = Workbook()

    # ═════════════════════════════ SERIE MUNICIPAL (datos) ══════════════════
    ws = wb.active
    ws.title = "Serie municipal"
    cols = ["codigo_dane", "municipio", "subregion", "periodo", "ano_inicio",
            "ano_fin", "clase", "hectareas", "hectareas_anuales", "fuente", "estimado"]
    head(ws, 1, ["Código DANE", "Municipio", "Territorial", "Periodo", "Año inicio",
                 "Año fin", "Clase", "Hectáreas", "Ha/año", "Fuente", "Estimado"],
         [12, 20, 12, 12, 10, 10, 18, 12, 12, 22, 10])
    for r, row in enumerate(serie[cols].itertuples(index=False), start=2):
        vals = list(row)
        vals[-1] = "Sí" if vals[-1] else "No"
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDE
            if c == 1:
                cell.number_format = "@"
            elif c in (5, 6):
                cell.number_format = "0"
            elif c in (8, 9):
                cell.number_format = NUM_HA1
            if vals[-1] == "Sí":
                cell.fill = FILL_EST
    ws.auto_filter.ref = f"A1:K{FIN}"
    ws.freeze_panes = "A2"

    D = "'Serie municipal'"
    RANGOS = {k: f"{D}!${c}$2:${c}${FIN}" for k, c in
              [("dane", "A"), ("mun", "B"), ("sub", "C"), ("per", "D"),
               ("clase", "G"), ("ha", "H"), ("hay", "I")]}

    def sumifs(criterios: list[tuple[str, str]], col="ha") -> str:
        crit = ",".join(f'{RANGOS[k]},{v}' for k, v in criterios)
        return f"=SUMIFS({RANGOS[col]},{crit})"

    # ═════════════════════════════ SERIE REGIONAL ═══════════════════════════
    ws = wb.create_sheet("Serie regional")
    head(ws, 1, ["Periodo", "Año inicio", "Año fin", "Años", "Deforestación (ha)",
                 "Ha/año", "Estimado", "Etiqueta"],
         [12, 10, 10, 8, 18, 12, 10, 12])
    for i, row in reg_d.iterrows():
        r = i + 2
        ws.cell(row=r, column=1, value=row["periodo"]).number_format = "@"
        ws.cell(row=r, column=2, value=int(row["ano_inicio"])).number_format = "0"
        ws.cell(row=r, column=3, value=int(row["ano_fin"])).number_format = "0"
        ws.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = "0"
        ws.cell(row=r, column=5,
                value=sumifs([("per", f"A{r}"), ("clase", '"Deforestación"')])
                ).number_format = NUM_HA
        ws.cell(row=r, column=6, value=f"=E{r}/D{r}").number_format = NUM_HA
        ws.cell(row=r, column=7, value="Sí" if row["estimado"] else "No")
        ws.cell(row=r, column=8, value=f'=A{r}&IF(G{r}="Sí","*","")')
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = BORDE
            if row["estimado"]:
                ws.cell(row=r, column=c).fill = FILL_EST
    nreg = len(reg_d) + 1
    tot = ws.cell(row=nreg + 1, column=4, value="TOTAL")
    tot.font = Font(bold=True)
    ws.cell(row=nreg + 1, column=5, value=f"=SUM(E2:E{nreg})").number_format = NUM_HA
    ws.cell(row=nreg + 1, column=5).font = Font(bold=True)
    ws.cell(row=nreg + 2, column=1,
            value="* Periodo estimado/calibrado (sin fuente municipal directa "
                  "en el paquete original). Ver Diccionario.").font = F_NOTA

    # ═════════════════════════════ AUX (cálculos) ═══════════════════════════
    ws = wb.create_sheet("Aux")
    ws.sheet_properties.tabColor = "9CA3AF"
    ws["A1"] = "Hoja auxiliar de cálculos (totales por municipio y ranking)"
    ws["A1"].font = F_H1
    head(ws, 3, ["Municipio", "Territorial", "Total 2000-2024 (ha)", "Ha/año prom.",
                 "", "Pos.", "Ranking municipio", "Total (ha)"],
         [22, 12, 20, 12, 3, 6, 22, 14])
    for i, m in enumerate(municipios):
        r = 4 + i
        ws.cell(row=r, column=1, value=m)
        sub = serie[serie["municipio"] == m]["subregion"].iloc[0]
        ws.cell(row=r, column=2, value=sub)
        ws.cell(row=r, column=3,
                value=sumifs([("mun", f"A{r}"), ("clase", '"Deforestación"')])
                ).number_format = NUM_HA1
        ws.cell(row=r, column=4, value=f"=C{r}/24").number_format = NUM_HA1
    fm, lm = 4, 3 + len(municipios)
    for k in range(10):
        r = 4 + k
        ws.cell(row=r, column=6, value=k + 1)
        ws.cell(row=r, column=8,
                value=f"=LARGE($C${fm}:$C${lm},F{r})").number_format = NUM_HA
        ws.cell(row=r, column=7,
                value=f"=INDEX($A${fm}:$A${lm},MATCH(H{r},$C${fm}:$C${lm},0))")
    # subregiones
    head(ws, 16, [], None)
    ws.cell(row=16, column=10, value="Territorial").font = F_HEAD
    ws.cell(row=16, column=10).fill = FILL_HEAD
    ws.cell(row=16, column=11, value="Total (ha)").font = F_HEAD
    ws.cell(row=16, column=11).fill = FILL_HEAD
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 14
    subregiones = ["Caribe", "Centro", "Atrato", "Nutibara", "Urrao"]
    for i, s in enumerate(subregiones):
        r = 17 + i
        ws.cell(row=r, column=10, value=s)
        ws.cell(row=r, column=11,
                value=sumifs([("sub", f"J{r}"), ("clase", '"Deforestación"')])
                ).number_format = NUM_HA

    # ═════════════════════════════ DASHBOARD ════════════════════════════════
    ws = wb.create_sheet("Dashboard", 1)
    ws.sheet_view.showGridLines = False
    ws["B2"] = "DASHBOARD — DEFORESTACIÓN CORPOURABA 2000-2024"
    ws["B2"].font = F_TITULO
    ws["B3"] = "19 municipios · 18 periodos de monitoreo · fórmulas auditables sobre la hoja «Serie municipal»"
    ws["B3"].font = F_SUB

    kpis = [
        ("Total deforestado (ha)", "='Serie regional'!E20", NUM_HA),
        ("Promedio anual (ha/año)", "='Serie regional'!E20/24", NUM_HA),
        ("Periodo más crítico",
         "=INDEX('Serie regional'!A2:A19,MATCH(MAX('Serie regional'!E2:E19),'Serie regional'!E2:E19,0))", "@"),
        ("Municipio más afectado", "=Aux!G4", "@"),
    ]
    for i, (lbl, formula, fmt) in enumerate(kpis):
        col = 2 + i * 3
        L = get_column_letter(col)
        ws.merge_cells(f"{L}5:{get_column_letter(col+1)}6")
        c = ws[f"{L}5"]
        c.value, c.number_format, c.font, c.alignment = formula, fmt, F_KPI_NUM, CENTRO
        for rr in (5, 6):
            for cc in range(col, col + 2):
                ws.cell(row=rr, column=cc).fill = FILL_KPI
        lb = ws[f"{L}7"]
        lb.value, lb.font = lbl, F_KPI_LBL
    ws["B8"] = "Nota: los periodos marcados con * son estimados/calibrados (2010-2012, 2018-2019, 2023-2024). 2015-2016 es dato real recuperado de la tabla municipal oficial."
    ws["B8"].font = F_NOTA

    lc = LineChart()
    lc.title = "Deforestación anualizada (ha/año) — jurisdicción CORPOURABA"
    lc.height, lc.width, lc.style = 9, 26, 12
    lc.y_axis.title = "ha/año"
    lc.add_data(Reference(wb["Serie regional"], min_col=6, min_row=1, max_row=nreg),
                titles_from_data=True)
    lc.set_categories(Reference(wb["Serie regional"], min_col=8, min_row=2, max_row=nreg))
    lc.series[0].smooth = False
    ws.add_chart(lc, "B10")

    bc = BarChart()
    bc.type, bc.title = "bar", "Top 10 municipios por deforestación acumulada (ha)"
    bc.height, bc.width, bc.style = 9, 13, 10
    bc.add_data(Reference(wb["Aux"], min_col=8, min_row=3, max_row=13),
                titles_from_data=True)
    bc.set_categories(Reference(wb["Aux"], min_col=7, min_row=4, max_row=13))
    bc.legend = None
    ws.add_chart(bc, "B30")

    pc = PieChart()
    pc.title = "Distribución por territorial (ha)"
    pc.height, pc.width = 9, 12
    pc.add_data(Reference(wb["Aux"], min_col=11, min_row=16, max_row=21),
                titles_from_data=True)
    pc.set_categories(Reference(wb["Aux"], min_col=10, min_row=17, max_row=21))
    ws.add_chart(pc, "J30")

    # ═════════════════════════════ CONSULTA MUNICIPIO ═══════════════════════
    ws = wb.create_sheet("Consulta municipio", 2)
    ws.sheet_view.showGridLines = False
    ws["B2"] = "CONSULTA POR MUNICIPIO"
    ws["B2"].font = F_TITULO
    ws["B4"] = "Seleccione municipio:"
    ws["B4"].font = F_H1
    sel = ws["E4"]
    sel.value = municipios[0]
    sel.font = Font(size=14, bold=True, color="1D4ED8")
    sel.fill = PatternFill("solid", start_color="FFFF00")
    sel.border = BORDE
    ws.column_dimensions["E"].width = 24
    dv = DataValidation(type="list", formula1=f"=Aux!$A$4:$A${lm}", allow_blank=False)
    dv.add(sel)
    ws.add_data_validation(dv)
    ws["G4"] = "(desplegable — celda amarilla)"
    ws["G4"].font = F_NOTA

    kpis_m = [
        ("Total 2000-2024 (ha)", 'SUMIFS({ha},{mun},$E$4,{clase},"Deforestación")', NUM_HA1),
        ("% del total regional", "B7/'Serie regional'!$E$20", NUM_PCT),
        ("Periodo más crítico", None, "@"),
        ("Territorial", "INDEX(Aux!$B$4:$B$22,MATCH($E$4,Aux!$A$4:$A$22,0))", "@"),
    ]
    for i, (lbl, f, fmt) in enumerate(kpis_m):
        col = 2 + i * 3
        L = get_column_letter(col)
        c = ws[f"{L}7"]
        if f:
            c.value = "=" + f.format(**RANGOS)
        c.number_format, c.font = fmt, Font(size=16, bold=True, color=VERDE)
        ws[f"{L}8"] = lbl
        ws[f"{L}8"].font = F_KPI_LBL

    head(ws, 10, ["Periodo", "Deforestación (ha)", "Ha/año", "Ha/año regional",
                  "Estimado"], [12, 18, 12, 16, 10])
    for i, p in enumerate(periodos):
        r = 11 + i
        rr = i + 2  # fila en Serie regional
        ws.cell(row=r, column=1, value=f"='Serie regional'!A{rr}").number_format = "@"
        ws.cell(row=r, column=2, value="=" +
                'SUMIFS({ha},{mun},$E$4,{clase},"Deforestación",{per},$A${r})'
                .replace("{r}", str(r)).format(**RANGOS, r=r)).number_format = NUM_HA1
        ws.cell(row=r, column=3,
                value=f"=B{r}/'Serie regional'!D{rr}").number_format = NUM_HA1
        ws.cell(row=r, column=4,
                value=f"='Serie regional'!F{rr}").number_format = NUM_HA
        ws.cell(row=r, column=5, value=f"='Serie regional'!G{rr}")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDE
    fin_m = 10 + len(periodos)
    ws["B7"].value = "=SUM(B11:B" + str(fin_m) + ")"
    ws["H7"] = (f"=INDEX(A11:A{fin_m},MATCH(MAX(B11:B{fin_m}),B11:B{fin_m},0))")
    ws["H7"].number_format = "@"
    ws["H7"].font = Font(size=16, bold=True, color=VERDE)

    lcm = LineChart()
    lcm.title = "Municipio seleccionado vs promedio regional (ha/año)"
    lcm.height, lcm.width, lcm.style = 10, 24, 12
    lcm.y_axis.title = "ha/año"
    lcm.add_data(Reference(ws, min_col=3, min_row=10, max_row=fin_m),
                 titles_from_data=True)
    lcm.add_data(Reference(ws, min_col=4, min_row=10, max_row=fin_m),
                 titles_from_data=True)
    lcm.set_categories(Reference(ws, min_col=1, min_row=11, max_row=fin_m))
    ws.add_chart(lcm, "G10")

    # ═════════════════════════════ MATRIZ ═══════════════════════════════════
    ws = wb.create_sheet("Matriz", 3)
    ws["A1"] = "Matriz de deforestación (ha) — municipio × periodo"
    ws["A1"].font = F_H1
    head(ws, 3, ["Municipio"] + periodos + ["TOTAL"],
         [20] + [11] * len(periodos) + [12])
    for i, m in enumerate(municipios):
        r = 4 + i
        ws.cell(row=r, column=1, value=m).border = BORDE
        for j in range(len(periodos)):
            c = j + 2
            L = get_column_letter(c)
            cell = ws.cell(row=r, column=c, value="=" +
                           'SUMIFS({ha},{mun},$A{r},{clase},"Deforestación",{per},{L}$3)'
                           .format(**RANGOS, r=r, L=L))
            cell.number_format = NUM_HA
            cell.border = BORDE
        tot_col = len(periodos) + 2
        tc = ws.cell(row=r, column=tot_col,
                     value=f"=SUM(B{r}:{get_column_letter(tot_col-1)}{r})")
        tc.number_format, tc.font, tc.border = NUM_HA, Font(bold=True), BORDE
    rt = 4 + len(municipios)
    ws.cell(row=rt, column=1, value="TOTAL").font = Font(bold=True)
    for c in range(2, len(periodos) + 3):
        L = get_column_letter(c)
        cell = ws.cell(row=rt, column=c, value=f"=SUM({L}4:{L}{rt-1})")
        cell.number_format, cell.font = NUM_HA, Font(bold=True)
    rango = f"B4:{get_column_letter(len(periodos)+1)}{rt-1}"
    ws.conditional_formatting.add(rango, ColorScaleRule(
        start_type="num", start_value=0, start_color="FFFFFF",
        mid_type="percentile", mid_value=70, mid_color="FDBA74",
        end_type="max", end_color="B91C1C"))
    ws.freeze_panes = "B4"
    ws.cell(row=rt + 2, column=1,
            value="Periodos 2010-2012, 2018-2019 y 2023-2024: valores "
                  "estimados/calibrados (2015-2016 es dato real). Escala de color blanco→rojo por intensidad."
            ).font = F_NOTA

    # ═════════════ HOJAS OPCIONALES (si la investigación ya corrió) ═════════
    extra = []
    din = PROC / "analisis" / "dinamica_bosque.csv"
    if din.exists():
        df = pd.read_csv(din, encoding="utf-8-sig")
        ws = wb.create_sheet("Dinámica bosque")
        head(ws, 1, [str(c) for c in df.columns], [16] * len(df.columns))
        for r, row in enumerate(df.itertuples(index=False), start=2):
            for c, v in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=v).border = BORDE
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
        ws.freeze_panes = "A2"
        extra.append("Dinámica bosque")
    hal = PROC / "analisis" / "hallazgos.json"
    if hal.exists():
        datos = json.loads(hal.read_text(encoding="utf-8"))
        ws = wb.create_sheet("Hallazgos")
        head(ws, 1, ["Tema", "Hallazgo", "Cifra", "Unidad", "Periodo", "Descripción"],
             [18, 45, 14, 12, 14, 90])
        for r, h in enumerate(datos, start=2):
            for c, k in enumerate(["tema", "titulo", "cifra", "unidad",
                                   "periodo_referencia", "descripcion"], start=1):
                ws.cell(row=r, column=c, value=h.get(k)).border = BORDE
            ws.cell(row=r, column=6).alignment = Alignment(wrap_text=True)
        ws.freeze_panes = "A2"
        extra.append("Hallazgos")

    # Hoja Territorios: deforestación por área protegida / POMCA / resguardo / consejo
    an = PROC / "analisis"
    bloques = []
    ap_csv = an / "areas_protegidas_serie.csv"
    if ap_csv.exists():
        ap = pd.read_csv(ap_csv, encoding="utf-8-sig")
        ap = ap[ap["clase"] == "Deforestación"]
        g = (ap.groupby(["nombre", "categoria"])["hectareas"].sum()
             .reset_index().sort_values("hectareas", ascending=False))
        bloques.append(("ÁREAS PROTEGIDAS (deforestación mapeada)",
                        ["Área protegida", "Categoría", "Deforestación (ha)"],
                        [[r["nombre"], r["categoria"], round(r["hectareas"], 1)]
                         for _, r in g.iterrows()]))
    pom_csv = an / "cartografia" / "pomcas_serie.csv"
    if pom_csv.exists():
        pom = pd.read_csv(pom_csv, encoding="utf-8-sig")
        g = (pom.groupby("pomca")["deforestacion_ha"].sum()
             .reset_index().sort_values("deforestacion_ha", ascending=False))
        bloques.append(("POMCAS — cuencas ordenadas (deforestación mapeada)",
                        ["POMCA", "", "Deforestación (ha)"],
                        [[r["pomca"], "", round(r["deforestacion_ha"], 1)]
                         for _, r in g.iterrows()]))
    ter_csv = an / "cartografia" / "territorios_oficiales.csv"
    if ter_csv.exists():
        ter = pd.read_csv(ter_csv, encoding="utf-8-sig")
        for tipo, titulo, sec in [("resguardo", "RESGUARDOS INDÍGENAS (límites oficiales)", "pueblo"),
                                  ("consejo_comunitario", "CONSEJOS COMUNITARIOS (límites oficiales)", "municipios")]:
            t = ter[ter["tipo"] == tipo]
            g = (t.groupby("nombre").agg(defo=("deforestacion_ha", "sum"),
                                         det=(sec, "first"))
                 .reset_index().sort_values("defo", ascending=False))
            bloques.append((titulo,
                            ["Territorio", "Pueblo / municipios", "Deforestación (ha)"],
                            [[r["nombre"], r["det"] if pd.notna(r["det"]) else "",
                              round(r["defo"], 1)] for _, r in g.iterrows()]))
    if bloques:
        ws = wb.create_sheet("Territorios")
        ws.column_dimensions["A"].width = 52
        ws.column_dimensions["B"].width = 34
        ws.column_dimensions["C"].width = 18
        ws["A1"] = "Deforestación por territorio (polígonos mapeados ≥1 ha, 2000-2024)"
        ws["A1"].font = F_H1
        fila = 3
        for titulo, cab, filas in bloques:
            ws.cell(row=fila, column=1, value=titulo).font = Font(bold=True, color=VERDE)
            fila += 1
            for c, t in enumerate(cab, start=1):
                celda = ws.cell(row=fila, column=c, value=t)
                celda.font, celda.fill, celda.border = F_HEAD, FILL_HEAD, BORDE
            fila += 1
            for datos_fila in filas:
                for c, v in enumerate(datos_fila, start=1):
                    celda = ws.cell(row=fila, column=c, value=v)
                    celda.border = BORDE
                    if c == 3:
                        celda.number_format = NUM_HA1
                fila += 1
            fila += 2  # espacio entre bloques
        extra.append("Territorios")

    # ═════════════════════════════ DICCIONARIO ══════════════════════════════
    # ═════════════════════════ TERRITORIALES CORPOURABA ═════════════════════
    ws = wb.create_sheet("Territoriales")
    ws.sheet_properties.tabColor = "1F7347"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "DEFORESTACIÓN POR TERRITORIAL — CORPOURABA 2000-2024"
    ws["B2"].font = F_TITULO
    ws["B3"] = ("Las 5 territoriales de la jurisdicción. Fórmulas auditables sobre la "
                "hoja «Serie municipal».")
    ws["B3"].font = F_SUB
    head(ws, 5, ["Territorial", "Deforestación total (ha)", "Ha/año",
                 "% del total", "Nº municipios"], [16, 22, 12, 12, 14])
    terr = ["Caribe", "Centro", "Atrato", "Nutibara", "Urrao"]
    f0, f1 = 6, 5 + len(terr)
    for i, t in enumerate(terr):
        r = 6 + i
        ws.cell(row=r, column=1, value=t).font = Font(bold=True)
        ws.cell(row=r, column=2,
                value=sumifs([("sub", f"$A{r}"), ("clase", '"Deforestación"')])
                ).number_format = NUM_HA1
        ws.cell(row=r, column=3, value=f"=B{r}/24").number_format = NUM_HA1
        ws.cell(row=r, column=4,
                value=f"=B{r}/SUM($B${f0}:$B${f1})").number_format = NUM_PCT
        ws.cell(row=r, column=5,
                value=f"=COUNTIF(Aux!$B$4:$B$22,$A{r})").number_format = "0"
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDE
    rt = f1 + 1
    ws.cell(row=rt, column=1, value="TOTAL").font = Font(bold=True)
    tot = ws.cell(row=rt, column=2, value=f"=SUM(B{f0}:B{f1})")
    tot.number_format, tot.font = NUM_HA1, Font(bold=True)
    ws.cell(row=rt, column=3, value=f"=B{rt}/24").number_format = NUM_HA1

    bc = BarChart()
    bc.type, bc.title = "bar", "Deforestación por territorial (ha)"
    bc.height, bc.width, bc.style = 8, 16, 11
    bc.add_data(Reference(ws, min_col=2, min_row=5, max_row=f1), titles_from_data=True)
    bc.set_categories(Reference(ws, min_col=1, min_row=f0, max_row=f1))
    bc.legend = None
    ws.add_chart(bc, "B14")
    ws.cell(row=rt + 2, column=1,
            value=("«Territorial» es la agrupación administrativa de CORPOURABA "
                   "(campo «subregion» en los datos). El % es sobre el total "
                   "deforestado de la jurisdicción (2000-2024).")).font = F_NOTA

    ws = wb.create_sheet("Diccionario")
    ws["A1"] = "Diccionario de datos y metodología"
    ws["A1"].font = F_H1
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100
    filas = [
        ("", ""),
        ("Fuente de datos", "Monitoreo de cambio de bosque CORPOURABA por periodos (shapefiles, Excel y rásters del paquete institucional 2000-2024), consolidado por el ETL del Observatorio."),
        ("Clases", "Bosque Estable (1) · Deforestación (2) · Sin Información (3) · Regeneración (4) · No Bosque Estable (5) — número = gridcode del paquete original."),
        ("hectareas", "Suma de hectáreas de la clase en el municipio y periodo (área geométrica en CRS métrico cuando hay shapefile)."),
        ("hectareas_anuales", "hectareas / años del periodo (los 5 primeros periodos duran 2 años) — usar SIEMPRE para comparar periodos."),
        ("fuente", "shapefile · excel · dbf-municipal · raster · estimado · estimado-calibrado-rat (mejor→peor)."),
        ("estimado", "Sí = el dato no proviene de una medición municipal directa y debe tratarse como referencia, no como cifra oficial."),
        ("", ""),
        ("Vacíos del paquete", "2010-2012: sin datos municipales (solo total departamental del ráster) · 2018-2019: carpeta inexistente · 2023-2024: shapefile municipal perdido. 2015-2016: la geometría (.shp) se perdió pero SÍ sobrevive la tabla de atributos municipal (.dbf) con área real, por lo que su cifra es dato medido."),
        ("Calibraciones", "2010-2012 = total departamental real del RAT × participación histórica de la jurisdicción (~18%), repartido por interpolación municipal. 2015-2016 = medición real desde Defor2015_2016_Mpios_Proj_Correg.dbf (área por municipio y clase de los 19 municipios; San Juan de Urabá con 0 ha)."),
        ("Control de calidad", "La serie consolidada difiere ≤0,3% de las hojas 'Cálculos' de los Excel institucionales originales en los 10 periodos comparables."),
        ("", ""),
        ("Fuentes por periodo", " · ".join(f"{p['id']}: {p['fuente']}" for p in meta["periodos"])),
    ]
    for i, (a, b) in enumerate(filas, start=2):
        ws.cell(row=i, column=1, value=a).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    # ═════════════════════════════ PORTADA ══════════════════════════════════
    ws = wb.create_sheet("Portada", 0)
    ws.sheet_view.showGridLines = False
    logo = ROOT / "frontend" / "public" / "logo-corpouraba.png"
    if logo.exists():
        try:
            from openpyxl.drawing.image import Image as XLImage

            imagen = XLImage(str(logo))
            imagen.width, imagen.height = 249, 100  # 75% del tamaño original
            ws.add_image(imagen, "F3")
        except ImportError:
            print("[aviso] Pillow no instalado: portada sin logo")
    for rango, texto, fuente in [
        ("B4", "OBSERVATORIO DE DEFORESTACIÓN", F_TITULO),
        ("B5", "Jurisdicción CORPOURABA · 19 municipios · 2000-2024", Font(size=16, color=VERDE)),
        ("B8", f"Fecha de generación: {date.today().strftime('%d/%m/%Y')}", F_SUB),
        ("B9", "Elaborado por: Alberto Vivas y Carlos Zuluaga", Font(size=12, bold=True)),
        ("B10", "Datos: paquete institucional de monitoreo de bosque CORPOURABA (IDEAM/SMByC)", F_SUB),
        ("B13", "Contenido del libro", F_H1),
    ]:
        ws[rango] = texto
        ws[rango].font = fuente
    hojas = [
        ("Dashboard", "KPIs, serie regional anualizada, top-10 municipios y distribución por territorial"),
        ("Territoriales", "Deforestación por cada una de las 5 territoriales de CORPOURABA (tabla + gráfico)"),
        ("Consulta municipio", "Ficha dinámica por municipio (lista desplegable en la celda amarilla)"),
        ("Matriz", "Municipio × periodo con escala de color y totales"),
        ("Serie regional", "Serie agregada por periodo (fórmulas sobre los datos)"),
        ("Serie municipal", f"Datos fuente completos ({n_datos} filas) con autofiltro"),
        ("Aux", "Cálculos auxiliares (totales por municipio, ranking, territoriales)"),
    ] + [(e, "Resultados de la investigación temática") for e in extra] + [
        ("Diccionario", "Definiciones, vacíos, calibraciones y control de calidad"),
    ]
    for i, (h, desc) in enumerate(hojas):
        r = 14 + i
        ws.cell(row=r, column=2, value=h).font = Font(bold=True, color=VERDE)
        ws.cell(row=r, column=3, value=desc).font = F_SUB
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 90
    nota = ws.cell(row=16 + len(hojas), column=2,
                   value="Todos los indicadores del libro son fórmulas de Excel (SUMIFS/INDEX/"
                         "MATCH/LARGE) sobre la hoja «Serie municipal»: puede auditarse y "
                         "extenderse sin herramientas adicionales. Los periodos con fondo "
                         "crema son estimados/calibrados.")
    nota.font = F_NOTA
    nota.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=16 + len(hojas), start_column=2,
                   end_row=17 + len(hojas), end_column=3)

    wb.save(OUT)
    print(f"OK -> {OUT}")
    print(f"Hojas: {wb.sheetnames}")
    print(f"Extras incluidos: {extra or 'ninguno (la investigación aún no genera salidas)'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
